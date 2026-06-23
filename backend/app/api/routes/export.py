"""Export service — generate PDF and Excel reports from metrics data."""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth import get_current_user
from app.database import get_db
from app.models import AdminUser

logger = logging.getLogger("export")
export_router = APIRouter(prefix="/export", tags=["Export"])


# ── Excel Export ───────────────────────────────────────────────────

@export_router.get("/excel/{device_id}")
async def export_excel(
    device_id: int,
    metric_name: Optional[str] = None,
    hours: int = Query(default=168, ge=1, le=720),  # default 1 week
    current_user: AdminUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export device metrics to Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="openpyxl not installed. Run: pip install openpyxl",
        )

    # Fetch metrics
    params: dict = {"device_id": device_id, "hours": hours}
    sql = """
        SELECT metric_name, time, metric_value, labels
        FROM device_metrics
        WHERE device_id = :device_id AND time > NOW() - make_interval(hours => :hours)
    """
    if metric_name:
        sql += " AND metric_name = :metric_name"
        params["metric_name"] = metric_name
    sql += " ORDER BY metric_name, time ASC"

    result = await db.execute(text(sql), params)
    rows = result.fetchall()

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for export")

    # Build Excel workbook
    wb = openpyxl.Workbook()

    # Group by metric_name
    metrics_data: dict[str, list] = {}
    for row in rows:
        metrics_data.setdefault(row.metric_name, []).append({
            "time": row.time,
            "value": row.metric_value,
            "labels": row.labels,
        })

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "NetMon Metrics Export"
    ws_summary["A1"].font = Font(size=16, bold=True)
    ws_summary["A2"] = f"Device ID: {device_id}"
    ws_summary["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_summary["A4"] = f"Period: Last {hours} hours"

    ws_summary["A6"] = "Metric"
    ws_summary["B6"] = "Data Points"
    ws_summary["C6"] = "Avg"
    ws_summary["D6"] = "Min"
    ws_summary["E6"] = "Max"
    for col in ["A", "B", "C", "D", "E"]:
        ws_summary[f"{col}6"].font = Font(bold=True)
        ws_summary[f"{col}6"].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        ws_summary[f"{col}6"].font = Font(bold=True, color="FFFFFF")

    row_idx = 7
    for mname, points in metrics_data.items():
        values = [p["value"] for p in points]
        ws_summary[f"A{row_idx}"] = mname
        ws_summary[f"B{row_idx}"] = len(points)
        ws_summary[f"C{row_idx}"] = round(sum(values) / len(values), 2) if values else 0
        ws_summary[f"D{row_idx}"] = round(min(values), 2) if values else 0
        ws_summary[f"E{row_idx}"] = round(max(values), 2) if values else 0
        row_idx += 1

    # Per-metric sheets with charts
    for mname, points in metrics_data.items():
        safe_name = mname[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=safe_name)

        ws["A1"] = "Time"
        ws["B1"] = mname
        for col in ["A", "B"]:
            ws[f"{col}1"].font = Font(bold=True)
            ws[f"{col}1"].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            ws[f"{col}1"].font = Font(bold=True, color="FFFFFF")

        for i, pt in enumerate(points, start=2):
            ws[f"A{i}"] = pt["time"].strftime("%Y-%m-%d %H:%M") if hasattr(pt["time"], "strftime") else str(pt["time"])
            ws[f"B{i}"] = round(pt["value"], 2)

        # Add chart
        if len(points) > 1:
            chart = LineChart()
            chart.title = mname
            chart.y_axis.title = "Value"
            chart.x_axis.title = "Time"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(points) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            ws.add_chart(chart, f"D2")

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"netmon_device_{device_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── PDF Export ─────────────────────────────────────────────────────

@export_router.get("/pdf/{device_id}")
async def export_pdf(
    device_id: int,
    metric_name: Optional[str] = None,
    hours: int = Query(default=168, ge=1, le=720),
    current_user: AdminUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Export device metrics to PDF report."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
        )
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.lineplots import LinePlot
        from reportlab.graphics.charts.textlabels import Label
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="reportlab not installed. Run: pip install reportlab",
        )

    # Fetch metrics
    params: dict = {"device_id": device_id, "hours": hours}
    sql = """
        SELECT metric_name, time, metric_value
        FROM device_metrics
        WHERE device_id = :device_id AND time > NOW() - make_interval(hours => :hours)
    """
    if metric_name:
        sql += " AND metric_name = :metric_name"
        params["metric_name"] = metric_name
    sql += " ORDER BY metric_name, time ASC"

    result = await db.execute(text(sql), params)
    rows = result.fetchall()

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for export")

    # Group data
    metrics_data: dict[str, list] = {}
    for row in rows:
        metrics_data.setdefault(row.metric_name, []).append({
            "time": row.time,
            "value": row.metric_value,
        })

    # Build PDF
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=18,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
    )

    elements = []

    # Title
    elements.append(Paragraph("NetMon — Metrics Report", title_style))
    elements.append(Paragraph(
        f"Device ID: {device_id} | Period: Last {hours}h | Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        subtitle_style,
    ))

    # Summary table
    summary_data = [["Metric", "Points", "Avg", "Min", "Max"]]
    for mname, points in metrics_data.items():
        values = [p["value"] for p in points]
        summary_data.append([
            mname,
            str(len(points)),
            f"{sum(values) / len(values):.2f}" if values else "0",
            f"{min(values):.2f}" if values else "0",
            f"{max(values):.2f}" if values else "0",
        ])

    summary_table = Table(summary_data, colWidths=[160, 60, 80, 80, 80])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3B82F6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("TOPPADDING", (0, 1), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Per-metric charts
    for mname, points in metrics_data.items():
        elements.append(Paragraph(f"<b>{mname}</b>", styles["Heading3"]))

        # Data table (last 20 points)
        recent = points[-20:]
        table_data = [["Time", "Value"]]
        for pt in recent:
            t = pt["time"].strftime("%Y-%m-%d %H:%M") if hasattr(pt["time"], "strftime") else str(pt["time"])
            table_data.append([t, f"{pt['value']:.2f}"])

        t = Table(table_data, colWidths=[200, 100])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#64748B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(t)

        # Simple line chart
        if len(points) > 2:
            drawing = Drawing(450, 150)
            lp = LinePlot()
            lp.x = 50
            lp.y = 30
            lp.width = 380
            lp.height = 100
            chart_data = [(i, p["value"]) for i, p in enumerate(points[-50:])]
            lp.data = [chart_data]
            lp.lines[0].strokeColor = colors.HexColor("#3B82F6")
            lp.lines[0].strokeWidth = 1.5
            drawing.add(lp)
            elements.append(drawing)

        elements.append(Spacer(1, 15))

    doc.build(elements)
    buf.seek(0)

    filename = f"netmon_device_{device_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
